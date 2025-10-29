# ==============================================================================
# --- VC Fund Model: Core Simulation Engine (v2.0.0) ---
# ==============================================================================
#
# v2.0.0: Enhanced annotations, improved print statements, and added verbose logging control
#
# ==============================================================================
import pandas as pd
import numpy as np
import logging
import heapq
from typing import List, Optional, Tuple, Dict, Any
from parameters import FundParameters, PortfolioResult, CompanyResult, Company
from utils import xirr
from waterfall import apply_fund_structure
import math as math

import waterfall

def round_to_hundred_thousand(number):
    """
    Round a number to the nearest 100,000 for simplified valuation modeling.
    
    Args:
        number: The number to round
        
    Returns:
        Number rounded to nearest 100,000
        
    Examples:
        round_to_hundred_thousand(6234589) -> 6200000
        round_to_hundred_thousand(6284589) -> 6300000
        round_to_hundred_thousand(49999) -> 0
        round_to_hundred_thousand(50000) -> 100000
    """
    return round(number / 100000) * 100000


def quick_simulate(stage: str, rng: np.random.Generator, from_valuation: float, n_simulations: int = 10000):
    """
    Monte Carlo simulation for company valuation progression between funding rounds.
    
    Uses fitted regression models to simulate realistic valuation multiples based on 
    current stage and valuation level. Higher current valuations tend to have lower 
    expected multiples (regression to the mean).
    
    Args:
        stage: Current funding stage ('Pre-Seed', 'Seed', 'Series A', 'Series B')
        rng: Random number generator for reproducible results
        from_valuation: Current company valuation to simulate from
        n_simulations: Number of simulation draws (default 10000)
        
    Returns:
        Array of simulated next-round valuations
    """
    # Regression model parameters fitted from real venture data
    model_params = {
        'Pre-Seed': {'alpha': 7.9438, 'beta': 0.4823, 'residual_std': 1.0643},
        'Seed': {'alpha': 9.7765, 'beta': 0.5139, 'residual_std': 0.7872},
        'Series A': {'alpha': 7.4050, 'beta': 0.3594, 'residual_std': 0.7012},
        'Series B': {'alpha': 6.7335, 'beta': 0.3115, 'residual_std': 0.8112}
    }
    
    if stage not in model_params:
        raise ValueError(f"Stage '{stage}' not supported")
    
    params = model_params[stage]
    
    # Calculate expected log(multiple) using regression model
    expected_log_multiple = params['alpha'] - params['beta'] * np.log(from_valuation)
    
    # Add random noise to simulate variation
    noise = rng.normal(0, params['residual_std'], n_simulations)
    log_multiples = expected_log_multiple + noise
    multiples = np.exp(log_multiples)
    
    # Return array of next-round valuations
    return from_valuation * multiples


def _trigger_capital_call(
    cash_on_hand: float,
    capital_called: float,
    params: FundParameters,
    current_time: float,
    gross_cash_flows: List[Tuple[float, float, int]],
    amount_needed: float = 0.0,
    verbose: bool = True
) -> Tuple[float, float]:
    """
    Fund treasury management: Calls capital from limited partners when needed.
    
    This function implements intelligent capital calling that:
    1. Maintains a minimum cash buffer for operations
    2. Calls capital in standard tranches to minimize LP disruption  
    3. Only calls when actually needed (not on a fixed schedule)
    4. Respects the total committed capital limit
    
    Args:
        cash_on_hand: Current fund cash balance
        capital_called: Total capital called to date
        params: Fund configuration parameters
        current_time: Current simulation time in months
        gross_cash_flows: Master transaction ledger
        amount_needed: Immediate cash requirement (e.g., for investment)
        verbose: Whether to print detailed logging information
        
    Returns:
        Tuple of (updated_cash_on_hand, updated_capital_called)
    """

    # Calculate minimum cash buffer as percentage of standard tranche size
    min_cash_balance = params.committed_capital * params.capital_calls.tranche_size_pct * params.capital_calls.minimum_cash_balance_pct

    # Determine if capital call is needed
    is_shortfall = amount_needed > cash_on_hand  # Can't cover immediate need
    would_be_below_buffer = (cash_on_hand - amount_needed) < min_cash_balance  # Would fall below safety buffer

    if is_shortfall or would_be_below_buffer:
        # Calculate precise amount needed to restore minimum buffer
        required_call = min_cash_balance - (cash_on_hand - amount_needed)
        targeted_call_amount = max(0, required_call)

        # Use standard tranche size to minimize LP disruption
        standard_tranche = params.committed_capital * params.capital_calls.tranche_size_pct
        desired_call_amount = max(targeted_call_amount, standard_tranche)

        # Constrain by remaining uncommitted capital
        call_amount_possible = params.committed_capital - capital_called
        amount_to_call = min(desired_call_amount, call_amount_possible)

        # Execute capital call if meaningful amount
        if amount_to_call > 1:
            cash_on_hand += amount_to_call
            capital_called += amount_to_call
            
            # Record capital call in transaction ledger (company_id = -2)
            gross_cash_flows.append((-amount_to_call, current_time, -2))
            
            if verbose:
                print(f"CAPITAL CALL EXECUTED: Called ${amount_to_call:,.0f} from LPs")
                print(f"  • Reason: {'Immediate shortfall' if is_shortfall else 'Buffer maintenance'}")
                print(f"  • New cash balance: ${cash_on_hand:,.0f}")
                print(f"  • Total capital called: ${capital_called:,.0f} ({capital_called/params.committed_capital:.1%} of commitment)")

    return cash_on_hand, capital_called


def _get_next_milestone_outcome(company: Company, params: FundParameters, rng: np.random.Generator) -> str:
    """
    Determines company fate at milestone events using probabilistic outcomes.
    
    At each milestone (typically 12-24 months), companies can:
    - Progress to next funding stage (if available and probable)
    - Exit via acquisition or IPO
    - Fail and shut down
    
    Probabilities are stage-specific and defined in fund parameters.
    
    Args:
        company: Company object with current stage information
        params: Fund parameters containing stage-specific probabilities
        rng: Random number generator for outcome selection
        
    Returns:
        Outcome string: 'exit', 'fail', or next stage name (e.g., 'Series A')
    """

    # Get probability rules for company's current stage
    stage_params = params.stages[company.current_stage]

    # Initialize basic outcomes always available
    choices = ['exit', 'fail']
    probabilities = [stage_params.prob_to_exit, stage_params.prob_to_fail]

    # Add progression option if next stage exists and is probable
    if stage_params.prob_to_next_stage is not None and stage_params.prob_to_next_stage > 0:
        current_stage_index = params.stages_order.index(company.current_stage)
        
        # Ensure we're not already in the final stage
        if current_stage_index < len(params.stages_order) - 1:
            next_stage_name = params.stages_order[current_stage_index + 1]
            choices.append(next_stage_name)
            probabilities.append(stage_params.prob_to_next_stage)

    # Normalize probabilities to ensure they sum to 1.0
    total_prob = sum(probabilities)
    
    if total_prob <= 0:
        logging.warning(f"Company {company.company_id} in stage '{company.current_stage}' has no valid outcome paths. Forcing failure.")
        return 'fail'

    normalized_probabilities = np.array(probabilities) / total_prob

    # Make probabilistic decision
    return rng.choice(choices, p=normalized_probabilities)


def _run_one_event_driven_simulation(params: FundParameters, rng: np.random.Generator, debug: bool = False, verbose: bool = True) -> Tuple[Optional[PortfolioResult], List[Any], List[Dict[str, Any]]]:
    """
    Core simulation engine: Models complete lifecycle of a single VC fund.
    
    This event-driven simulation processes the fund from first investment to final 
    dissolution, handling:
    - Investment sourcing and execution
    - Company milestone progression
    - Follow-on investment decisions  
    - Portfolio company exits and failures
    - Management fee payments
    - Capital calling and cash management
    
    Args:
        params: Complete fund configuration parameters
        rng: Random number generator for reproducible results
        debug: Enable detailed debugging output (currently unused)
        verbose: Control print statement output for clean batch runs
        
    Returns:
        Tuple of (PortfolioResult, gross_cash_flows, waterfall_details, net_lp_flows, debug_log)
    """
    if verbose:
        print(f'Starting new fund simulation with RNG seed: {hash(str(rng)) % 1000000}')

    # Initialize fund state variables
    time, cash_on_hand, capital_called = 0.0, 0.0, 0.0
    portfolio: Dict[int, Company] = {}
    gross_cash_flows: List[Tuple[float, float, int]] = []
    final_company_results = []

    # Investment pacing and strategy tracking
    deals_per_year = {y: 0 for y in range(1, int(params.investment_period_months / 12) + 2)}
    investments_made, capital_constrained_flag, strategy_has_pivoted = 0, False, False
    cumulative_investment_dollars = 0
    recycled_capital_total = 0.0
    cash_back_tracker = 0

    debug_log: List[Dict[str, Any]] = []

    # Initialize follow-on strategy based on fund type
    strategy_type = params.follow_on_strategy.type
    if strategy_type == "spray_and_pray":
        current_pro_rata_rate = 0.0
    elif strategy_type == "passive":
        current_pro_rata_rate = params.follow_on_strategy.passive_participation_rate
    elif strategy_type == "pro_rata":
        current_pro_rata_rate = 1.0
    else:  # 'dynamic'
        current_pro_rata_rate = 1.0

    # Determine actual fund lifespan including potential extensions
    actual_fund_lifespan = params.fund_lifespan_months
    extensions_granted = 0
    
    for extension_prob in params.prob_of_extensions:
        if rng.random() < extension_prob:
            actual_fund_lifespan += 12
            extensions_granted += 1
        else:
            break

    if verbose and extensions_granted > 0:
        print(f"FUND SETUP: Granted {extensions_granted} extension(s), total lifespan: {actual_fund_lifespan/12:.1f} years")

    # Initialize event-driven simulation queue
    event_queue: List[Tuple[float, str, dict]] = []

    # Bootstrap simulation with initial events
    cash_on_hand, capital_called = _trigger_capital_call(cash_on_hand, capital_called, params, time, gross_cash_flows, verbose=verbose)
    
    # Schedule first investment consideration
    heapq.heappush(event_queue, (rng.uniform(0, 6), "CONSIDER_NEW_INVESTMENT", {}))

    # Schedule management fee payments for entire fund life
    for year in range(1, int(np.ceil(actual_fund_lifespan / 12)) + 1):
        heapq.heappush(event_queue, (year * 12, "FEE_PAYMENT", {}))

    # Schedule strategy review if using dynamic approach
    if strategy_type == "dynamic":
        heapq.heappush(event_queue, (params.follow_on_strategy.strategy_review_month, "STRATEGY_REVIEW", {}))

    step_counter = 0

    # Main event processing loop
    while event_queue:
        step_counter += 1
        time, event_type, data = heapq.heappop(event_queue)

        # Skip events beyond fund lifespan
        if time > actual_fund_lifespan: 
            continue

        if verbose:
            print(f"\n{'='*60}")
            print(f"STEP {step_counter}: {event_type} at {time:.1f} months ({time/12:.1f} years)")
            print(f"Fund Status: ${cash_on_hand:,.0f} cash, {len(portfolio)} active companies")
            print(f"{'='*60}")

        # Process management fee payments
        if event_type == "FEE_PAYMENT":
            if verbose:
                print("\n--- ANNUAL MANAGEMENT FEE PROCESSING ---")
            
            # Determine fee calculation basis
            is_commitment_period = time <= params.investment_period_months
            
            if is_commitment_period:
                fee_base = params.committed_capital
                fee_rate = params.mgmt_fee_commitment_period_rate
                fee_period = "commitment period"
            else:
                # Post-commitment: fee based on invested capital in active companies
                fee_base = sum(c.total_invested for c in portfolio.values() 
                              if c.status in ["active_supported", "active_passive"])
                fee_rate = params.mgmt_fee_post_commitment_period_rate
                fee_period = "post-commitment period"

            # Handle fund wind-down scenario
            if time >= params.investment_period_months and fee_base == 0 and cash_back_tracker == 0:
                if verbose:
                    print(f"FUND WIND-DOWN: No remaining assets under management")
                    print(f"Returning unused cash position of ${cash_on_hand:,.0f} to limited partners")
                
                gross_cash_flows.append((cash_on_hand, time, 9999))
                cash_on_hand = 0
                cash_back_tracker = 1
                break

            fee_amount = fee_base * fee_rate

            if verbose:
                print(f"Fee calculation ({fee_period}):")
                print(f"  • Fee base (assets under management): ${fee_base:,.0f}")
                print(f"  • Annual fee rate: {fee_rate:.1%}")
                print(f"  • Annual fee amount: ${fee_amount:,.0f}")

            # Ensure sufficient cash for fee payment
            cash_on_hand, capital_called = _trigger_capital_call(cash_on_hand, capital_called, params, time, gross_cash_flows, amount_needed=fee_amount, verbose=verbose)

            # Process fee payment
            if cash_on_hand >= fee_amount:
                cash_on_hand -= fee_amount
                gross_cash_flows.append((-fee_amount, time, -1))  # company_id = -1 for fees
                
                if verbose:
                    print(f"FEE PAID: ${fee_amount:,.0f} management fee processed")
                    print(f"Remaining cash balance: ${cash_on_hand:,.0f}")
            else:
                if verbose:
                    print(f"WARNING: Insufficient cash for ${fee_amount:,.0f} fee payment - fund impaired")
                capital_constrained_flag = True

        # Process new investment considerations
        elif event_type == "CONSIDER_NEW_INVESTMENT":
            if verbose:
                print("\n--- NEW INVESTMENT OPPORTUNITY EVALUATION ---")

            current_year = int(time / 12) + 1
            
            # Investment eligibility gates
            is_in_investment_period = time <= params.investment_period_months
            has_deal_capacity = deals_per_year.get(current_year, 0) < params.max_deals_per_year
            has_portfolio_capacity = investments_made < params.num_investments

            if verbose:
                print(f"Investment eligibility check:")
                print(f"  • In investment period: {'✓' if is_in_investment_period else '✗'} ({time:.1f}/{params.investment_period_months} months)")
                print(f"  • Annual deal capacity: {'✓' if has_deal_capacity else '✗'} ({deals_per_year.get(current_year, 0)}/{params.max_deals_per_year} deals in year {current_year})")
                print(f"  • Portfolio capacity: {'✓' if has_portfolio_capacity else '✗'} ({investments_made}/{params.num_investments} companies)")

            if is_in_investment_period and has_deal_capacity and has_portfolio_capacity:
                if verbose:
                    print("\n--- DEAL STRUCTURING ---")

                # Determine investment stage based on dynamic allocation
                year_allocations = next((item.allocation for item in params.dynamic_stage_allocation 
                                       if item.year == current_year), None)
                if not year_allocations:
                    year_allocations = params.dynamic_stage_allocation[-1].allocation

                stages = list(year_allocations.keys())
                alloc_probs = list(year_allocations.values())
                chosen_stage = rng.choice(stages, p=alloc_probs)

                # Generate company valuation and investment terms
                dist = params.stages[chosen_stage].post_money_valuation_dist
                post_money_valuation = np.minimum(params.stages[chosen_stage].max_valuation,np.maximum(params.stages[chosen_stage].min_valuation, round_to_hundred_thousand(
                    rng.lognormal(mean=dist.mu_log, sigma=dist.sigma_log))))

                ownership_target = params.initial_ownership_targets[chosen_stage]
                initial_investment_amount = post_money_valuation * ownership_target

                if verbose:
                    print(f"Deal terms generated:")
                    print(f"  • Stage: {chosen_stage}")
                    print(f"  • Post-money valuation: ${post_money_valuation:,.0f}")
                    print(f"  • Target ownership: {ownership_target:.1%}")
                    print(f"  • Required investment: ${initial_investment_amount:,.0f}")

                # Investment budget and execution checks
                investable_capital_total = params.committed_capital * params.target_investable_capital_pct
                investable_capital_remaining = investable_capital_total - cumulative_investment_dollars

                if verbose:
                    print(f"\nInvestment budget analysis:")
                    print(f"  • Total investable capital: ${investable_capital_total:,.0f}")
                    print(f"  • Already invested: ${cumulative_investment_dollars:,.0f}")
                    print(f"  • Remaining budget: ${investable_capital_remaining:,.0f}")

                if investable_capital_remaining >= initial_investment_amount:
                    # Ensure sufficient cash availability
                    cash_on_hand, capital_called = _trigger_capital_call(cash_on_hand, capital_called, params, time, gross_cash_flows, amount_needed=initial_investment_amount, verbose=verbose)

                    if cash_on_hand >= initial_investment_amount:
                        # Execute investment
                        cash_on_hand -= initial_investment_amount
                        investments_made += 1
                        cumulative_investment_dollars += initial_investment_amount
                        deals_per_year[current_year] += 1
                        new_company_id = investments_made

                        # Create portfolio company
                        new_company = Company(
                            company_id=new_company_id,
                            start_time=time,
                            initial_investment=initial_investment_amount,
                            current_stage=chosen_stage,
                            ownership=initial_investment_amount/post_money_valuation,
                            valuation=post_money_valuation
                        )
                        portfolio[new_company_id] = new_company

                        # Record investment transaction
                        gross_cash_flows.append((-initial_investment_amount, time, new_company_id))

                        # Schedule first company milestone
                        time_to_milestone = params.stages[chosen_stage].time_in_stage_months
                        heapq.heappush(event_queue, (time + time_to_milestone, "MILESTONE", {"id": new_company_id}))

                        if verbose:
                            print(f"\n✓ INVESTMENT EXECUTED:")
                            print(f"  • Company {new_company_id} added to portfolio")
                            print(f"  • Stage: {chosen_stage}")
                            print(f"  • Investment: ${initial_investment_amount:,.0f}")
                            print(f"  • Ownership: {ownership_target:.1%}")
                            print(f"  • Next milestone: {time + time_to_milestone:.1f} months")
                            print(f"  • Remaining cash: ${cash_on_hand:,.0f}")
                            print(f"  • Remaining budget: ${investable_capital_total - cumulative_investment_dollars:,.0f}")

                    else:
                        if verbose:
                            print(f"✗ INVESTMENT BLOCKED: Insufficient cash after capital call attempt")
                else:
                    if verbose:
                        print(f"✗ INVESTMENT BLOCKED: Exceeds remaining investment budget")

            # Schedule next investment consideration
            if is_in_investment_period:
                time_to_next_deal = (12 / params.max_deals_per_year) * (0.5 + rng.random())
                heapq.heappush(event_queue, (time + time_to_next_deal, "CONSIDER_NEW_INVESTMENT", {}))

        # Process company milestone events
        elif event_type == "MILESTONE":
            company_id = data["id"]
            
            if company_id not in portfolio:
                if verbose:
                    print(f"Company {company_id} not found in portfolio - already exited")
                continue

            
            company = portfolio[company_id]
            outcome = _get_next_milestone_outcome(company, params, rng)
            #print(outcome)

            if verbose:
                print(f"\n--- MILESTONE EVENT: COMPANY {company_id} ---")
                print(f"Current status: {company.current_stage} stage, ${company.valuation:,.0f} valuation")
                print(f"Milestone outcome: {outcome.upper()}")
            
            if outcome == 'exit':

                exit_stage_name = params.stages_order[params.stages_order.index(company.current_stage) + 1]
                #print(f"Exit stage name:{exit_stage_name}")
                # Calculate exit proceeds using valuation simulation
                exit_valuation = round_to_hundred_thousand(np.minimum(params.stages[exit_stage_name].max_valuation,np.maximum(params.stages[exit_stage_name].min_valuation,
                    quick_simulate(stage=company.current_stage, rng=rng, 
                                 from_valuation=company.valuation, n_simulations=1)[0])))
                
                exit_proceeds = exit_valuation * company.ownership
                valuation_multiple = exit_valuation / company.valuation

                if verbose:
                    print(f"EXIT TRANSACTION:")
                    print(f"  • Entry valuation: ${company.valuation:,.0f}")
                    print(f"  • Exit valuation: ${exit_valuation:,.0f}")
                    print(f"  • Valuation multiple: {valuation_multiple:.1f}x")
                    print(f"  • Fund ownership: {company.ownership:.1%}")
                    print(f"  • Gross proceeds: ${exit_proceeds:,.0f}")

                # Handle capital recycling if enabled
                if params.allow_recycling:
                    recycling_limit = params.committed_capital * params.recycling_limit_pct_of_commitment
                    recycling_room = max(0, recycling_limit - recycled_capital_total)
                    amount_to_recycle = min(exit_proceeds, recycling_room)
                    
                    if amount_to_recycle > 0:
                        cash_on_hand += amount_to_recycle
                        recycled_capital_total += amount_to_recycle
                        proceeds_for_distribution = exit_proceeds - amount_to_recycle
                        
                        if proceeds_for_distribution > 0:
                            gross_cash_flows.append((proceeds_for_distribution, time, company_id))
                            
                        if verbose:
                            print(f"CAPITAL RECYCLING:")
                            print(f"  • Recycled for new investments: ${amount_to_recycle:,.0f}")
                            print(f"  • Distributed to LPs: ${proceeds_for_distribution:,.0f}")
                            print(f"  • Total recycled to date: ${recycled_capital_total:,.0f}")
                    else:
                        gross_cash_flows.append((exit_proceeds, time, company_id))
                        if verbose:
                            print(f"All ${exit_proceeds:,.0f} proceeds distributed (recycling limit reached)")
                else:
                    gross_cash_flows.append((exit_proceeds, time, company_id))
                    if verbose:
                        print(f"All ${exit_proceeds:,.0f} proceeds distributed (recycling disabled)")

                # Finalize company and remove from active portfolio
                company.finalize('exited', time=time, proceeds=exit_proceeds, exit_valuation=exit_valuation)
                final_company_results.append(company.generate_result())
                del portfolio[company_id]

            elif outcome == 'fail':
                if verbose:
                    print(f"COMPANY FAILURE: Company {company_id} shut down with zero recovery")
                
                company.finalize('failed', time=time, proceeds=0, exit_valuation=0, reason='probabilistic_failure')
                final_company_results.append(company.generate_result())
                del portfolio[company_id]

            else:  # Company progresses to next stage
                next_stage_name = outcome
                
                # Simulate new valuation using progression model
                new_post_money = round_to_hundred_thousand(np.minimum(params.stages[outcome].max_valuation,np.maximum(params.stages[outcome].min_valuation,
                    quick_simulate(stage=company.current_stage, rng=rng, 
                                 from_valuation=company.valuation, n_simulations=1)[0])))
                
                valuation_multiple = new_post_money / company.valuation
                target_dilution = params.stages[company.current_stage].target_dilution_pct
                
                # Calculate round economics
                new_pre_money = new_post_money * (1 - target_dilution) if target_dilution < 1 else new_post_money
                total_round_size = new_post_money - new_pre_money if target_dilution < 1 else 0
                new_valuation = new_pre_money + total_round_size
                round_dilution = total_round_size / new_valuation if new_valuation > 0 else 0

                if verbose:
                    print(f"FUNDING ROUND PROGRESSION:")
                    print(f"  • Advancing to: {next_stage_name}")
                    print(f"  • Previous valuation: ${company.valuation:,.0f}")
                    print(f"  • Valuation multiple: {valuation_multiple:.1f}x")
                    print(f"  • New pre-money: ${new_pre_money:,.0f}")
                    print(f"  • Total round size: ${total_round_size:,.0f}")
                    print(f"  • New post-money: ${new_valuation:,.0f}")
                    print(f"  • Round dilution: {round_dilution:.1%}")

                # Handle follow-on investment decisions
                if company.status == "active_supported":
                    ownership_before_round = company.ownership
                    pro_rata_right = ownership_before_round * total_round_size
                    follow_on_amount = pro_rata_right * current_pro_rata_rate

                    if verbose:
                        print(f"\nFOLLOW-ON DECISION:")
                        print(f"  • Current ownership: {ownership_before_round:.1%}")
                        print(f"  • Pro-rata right: ${pro_rata_right:,.0f}")
                        print(f"  • Fund's pro-rata rate: {current_pro_rata_rate:.1%}")
                        print(f"  • Proposed follow-on: ${follow_on_amount:,.0f}")

                    will_invest = current_pro_rata_rate > 0

                    if will_invest and follow_on_amount > 0:
                        investable_capital_remaining = (params.committed_capital * params.target_investable_capital_pct) - cumulative_investment_dollars

                        if investable_capital_remaining >= follow_on_amount:
                            cash_on_hand, capital_called = _trigger_capital_call(cash_on_hand, capital_called, params, time, gross_cash_flows, amount_needed=follow_on_amount, verbose=verbose)
                            
                            if cash_on_hand >= follow_on_amount:
                                # Execute follow-on investment
                                cash_on_hand -= follow_on_amount
                                cumulative_investment_dollars += follow_on_amount
                                gross_cash_flows.append((-follow_on_amount, time, company_id))
                                
                                new_ownership = (ownership_before_round * new_pre_money + follow_on_amount) / new_valuation
                                company.add_follow_on(amount=follow_on_amount, new_stage=next_stage_name, 
                                                    time=time, premoney_valuation=new_pre_money, 
                                                    round_dilution=round_dilution, new_valuation=new_valuation, 
                                                    new_ownership=new_ownership)
                                
                                if verbose:
                                    print(f"✓ FOLLOW-ON EXECUTED: ${follow_on_amount:,.0f} invested")
                                    print(f"  • New ownership: {new_ownership:.1%}")
                                    print(f"  • Remaining budget: ${(params.committed_capital * params.target_investable_capital_pct) - cumulative_investment_dollars:,.0f}")
                            else:
                                # Pass due to insufficient cash
                                new_ownership = (ownership_before_round * new_pre_money) / new_valuation
                                company.pass_on_round(next_stage_name, time, new_pre_money, round_dilution, new_valuation, new_ownership)
                            if verbose:
                                print(f"✗ FOLLOW-ON DECLINED: Exceeds investment budget")
                        else:
                            # Pass due to strategy
                            new_ownership = (ownership_before_round * new_pre_money) / new_valuation
                            company.pass_on_round(next_stage_name, time, new_pre_money, round_dilution, new_valuation, new_ownership)
                            if verbose:
                                print(f"✗ FOLLOW-ON DECLINED: Exceeds investment budget")
                    else:
                        # Pass due to strategy
                        new_ownership = (ownership_before_round * new_pre_money) / new_valuation
                        company.pass_on_round(next_stage_name, time, new_pre_money, round_dilution, new_valuation, new_ownership)
                        if verbose:
                            print(f"✗ FOLLOW-ON DECLINED: Pass due to strategy")        
                else:  # company.status == "active_passive"
                    # Passive companies: update ownership due to dilution only
                    ownership_before_round = company.ownership
                    new_ownership = (ownership_before_round * new_pre_money) / new_valuation
                    company.pass_on_round(next_stage_name, time, new_pre_money, round_dilution, new_valuation, new_ownership)
                    
                    if verbose:
                        print(f"PASSIVE DILUTION: Ownership updated from {ownership_before_round:.1%} to {new_ownership:.1%}")

                # Schedule next milestone for progressing company
                time_to_milestone = params.stages[next_stage_name].time_in_stage_months
                heapq.heappush(event_queue, (time + time_to_milestone, "MILESTONE", {"id": company_id}))
                
                if verbose:
                    print(f"Next milestone scheduled: {time + time_to_milestone:.1f} months")

        elif event_type == "STRATEGY_REVIEW":
            # Placeholder for dynamic strategy adjustment logic
            if verbose:
                print(f"\n--- STRATEGY REVIEW ---")
                print(f"Dynamic strategy review at {time:.1f} months (placeholder)")

    # --- Fund Wind-down and Final Accounting ---
    if verbose:
        print(f"\n{'='*60}")
        print(f"FUND LIFECYCLE COMPLETE: {actual_fund_lifespan/12:.1f} years")
        print(f"{'='*60}")

    # Handle remaining portfolio companies at fund termination
    for company in portfolio.values():
        if company.status in ["active_supported", "active_passive"]:
            company.timeout(time=actual_fund_lifespan)
            final_company_results.append(company.generate_result())
            if verbose:
                print(f"Company {company.company_id} timed out - marked as failed at fund termination")

    if cash_on_hand > 0:
        gross_cash_flows.append((cash_on_hand, time, 9999))
        cash_on_hand -= cash_on_hand

    # Apply fund structure (waterfall) to calculate net LP distributions
    net_lp_flows, final_fund_life, waterfall_details = apply_fund_structure(gross_cash_flows, params, actual_fund_lifespan, verbose)

    # Calculate performance metrics
    gross_flows_for_irr = [(amt, t) for amt, t, cid in gross_cash_flows if (cid >= 0 and cid < 9999)]
    
    total_invested = sum(abs(cf[0]) for cf in gross_flows_for_irr if cf[0] < 0)
    total_returned = sum(cf[0] for cf in gross_flows_for_irr if cf[0] > 0)
    gross_multiple = total_returned / total_invested if total_invested > 0 else 0
    gross_irr = xirr(gross_flows_for_irr, time_unit='months')

    net_lp_flows_list = [(row['amount'], row['time_months']) for _, row in net_lp_flows.iterrows()]
    net_invested = sum(abs(cf[0]) for cf in net_lp_flows_list if cf[0] < 0)
    net_returned = sum(cf[0] for cf in net_lp_flows_list if cf[0] > 0)
    net_multiple = net_returned / net_invested if net_invested > 0 else 0
    net_irr = xirr(net_lp_flows_list, time_unit='months')

    # Calculate operational metrics
    investment_flows = [abs(cf[0]) for cf in gross_flows_for_irr if cf[0] < 0]
    average_check_size = sum(investment_flows) / len(investment_flows) if investment_flows else 0.0

    # Handle IRR calculation failures
    if net_irr is None: net_irr = -1.0
    if gross_irr is None: gross_irr = -1.0

    if verbose:
        print(f"\n--- FINAL PERFORMANCE SUMMARY ---")
        print(f"Portfolio companies: {len(final_company_results)}")
        print(f"Total invested: ${total_invested:,.0f}")
        print(f"Total returned: ${total_returned:,.0f}")
        print(f"Gross multiple: {gross_multiple:.2f}x")
        print(f"Net multiple (to LPs): {net_multiple:.2f}x")
        print(f"Gross IRR: {gross_irr:.1%}")
        print(f"Net IRR (to LPs): {net_irr:.1%}")
        print(f"Fund extensions used: {extensions_granted}")
        print(f"Average check size: ${average_check_size:,.0f}")
        if capital_constrained_flag:
            print(f"⚠️  Fund experienced capital constraints")

    # Package final results
    result = PortfolioResult(
        gross_irr=gross_irr,
        net_irr=net_irr,
        gross_multiple=gross_multiple,
        net_multiple=net_multiple,
        capital_constrained=capital_constrained_flag,
        final_fund_life_years=final_fund_life,
        company_results=final_company_results,
        num_extensions=extensions_granted,
        average_check_size=average_check_size
    )

    return result, gross_cash_flows, waterfall_details, net_lp_flows, debug_log


def run_monte_carlo(params: FundParameters, num_simulations: int, seed: Optional[int] = None, verbose: bool = False) -> List[PortfolioResult]:
    """
    Orchestrates Monte Carlo simulation of VC fund performance.
    
    Runs multiple independent fund lifecycle simulations to generate statistical
    distributions of performance outcomes. Each simulation uses independent random
    draws while maintaining reproducibility through controlled seeding.
    
    Args:
        params: Complete fund configuration parameters
        num_simulations: Number of independent fund simulations to run
        seed: Random seed for reproducible results (None for random)
        verbose: Enable detailed logging for individual simulations
        
    Returns:
        Tuple of (results_list, gross_flows_list, waterfall_log, net_lp_flows_log)
        where results_list contains PortfolioResult objects for each simulation
    """

    # Initialize master random number generator
    rng = np.random.default_rng(seed)
    results: List[PortfolioResult] = []
    waterfall_details_list: List[pd.DataFrame] = []
    net_lp_flows_list: List[pd.DataFrame] = []
    gross_flows = []

    logging.info(f"Starting Monte Carlo simulation: {num_simulations} runs with seed={seed}")
    
    if not verbose:
        print(f"Running {num_simulations} fund simulations...")

    # Execute simulation runs
    for i in range(num_simulations):
        if not verbose:
            # Show progress for batch runs
            if (num_simulations >= 100) and ((i + 1) % (num_simulations // 10) == 0):
                print(f"  Progress: {i+1}/{num_simulations} ({(i+1)/num_simulations:.0%}) complete")
        else:
            print(f"\n{'#'*80}")
            print(f"STARTING SIMULATION {i+1} OF {num_simulations}")
            print(f"{'#'*80}")

        # Create independent RNG for this simulation
        sim_rng = np.random.default_rng(rng.integers(1e9))

        # Run single simulation
        result, gross_cash_flows, waterfall_details, net_lp_flows_dataframe, _ = _run_one_event_driven_simulation(params, sim_rng, debug=False, verbose=verbose)

        # Store results with simulation tracking
        if result:
            results.append(result)
            gross_flows.append(gross_cash_flows)
            
            # Add simulation number to tracking dataframes
            waterfall_details['simulation_number'] = i + 1
            waterfall_details_list.append(waterfall_details)

            net_lp_flows_dataframe['simulation_number'] = i + 1
            net_lp_flows_list.append(net_lp_flows_dataframe)

    # Consolidate tracking data
    waterfall_log = pd.concat(waterfall_details_list, ignore_index=True) if waterfall_details_list else pd.DataFrame()
    net_lp_flow_log = pd.concat(net_lp_flows_list, ignore_index=True) if net_lp_flows_list else pd.DataFrame()

    print(f"Monte Carlo simulation complete: {len(results)} successful runs")
    logging.info(f"Monte Carlo simulation complete: {len(results)} successful runs")

    return results, gross_flows, waterfall_log, net_lp_flow_log


def debug_one_simulation(params: FundParameters, rng: np.random.Generator, verbose: bool = True) -> Tuple[Optional[PortfolioResult], List, List[Dict[str, Any]]]:
    """
    Runs a single fund simulation with full debugging output and data retention.
    
    Unlike batch Monte Carlo runs, this function preserves all intermediate data
    including detailed cash flows and debug logs for deep analysis of fund mechanics.
    
    Args:
        params: Fund configuration parameters
        rng: Pre-configured random number generator  
        verbose: Enable detailed step-by-step logging
        
    Returns:
        Tuple of (PortfolioResult, cash_flows_list, debug_log_list)
        with complete transaction history and debugging information
    """

    print("Running single simulation in debug mode...")
    result, cash_flows, waterfall_details, net_lp_flows, debug_log = _run_one_event_driven_simulation(params, rng, debug=True, verbose=verbose)
    
    if result:
        print(f"\nDebug simulation complete:")
        print(f"  • Performance: {result.gross_multiple:.2f}x gross multiple, {result.gross_irr:.1%} gross IRR")
        print(f"  • Portfolio: {len(result.company_results)} companies")
        print(f"  • Transactions: {len(cash_flows)} cash flow events")
        
    return result, cash_flows, debug_log


def convert_multiple_simulations_to_excel_with_flows(all_results, all_gross_flows, waterfall_log=None, net_lp_flows_log=None, filename="multi_simulation_fund_performance.xlsx"):
    """
    Export comprehensive Monte Carlo simulation results to Excel workbook.
    
    Creates multi-sheet Excel file containing:
    - Company-level performance summary across all simulations
    - Round-by-round investment history for each company  
    - Complete cash flow ledger with transaction details
    - Waterfall distribution mechanics (if provided)
    - Net LP cash flows after fees and carry (if provided)
    
    Args:
        all_results: List of PortfolioResult objects from Monte Carlo runs
        all_gross_flows: List of gross cash flow ledgers for each simulation
        waterfall_log: Optional DataFrame with waterfall calculation details
        net_lp_flows_log: Optional DataFrame with net LP distribution flows
        filename: Output Excel file name
        
    Returns:
        Tuple of (companies_df, history_df, flows_df, waterfall_log, net_lp_flows_log)
    """
    
    from openpyxl.utils import get_column_letter
    
    all_company_data = []
    all_investment_history = []
    all_cash_flows_data = []
    
    print(f"Processing {len(all_results)} simulation results for Excel export...")
    
    # Process each simulation's results
    for sim_num, (result, gross_flows) in enumerate(zip(all_results, all_gross_flows), 1):
        company_results = result.company_results
        
        # Extract company-level summary data
        for comp_result in company_results:
            all_company_data.append({
                'Simulation_Number': sim_num,
                'Company_ID': comp_result.company_id,
                'Outcome': comp_result.outcome,
                'Failure_Reason': comp_result.failure_reason if comp_result.failure_reason else '',
                'Time_to_Exit_Months': comp_result.time_to_exit_months,
                'Time_to_Exit_Years': comp_result.time_to_exit_months / 12,
                'Total_Invested': comp_result.total_invested,
                'Exit_Proceeds': comp_result.exit_proceeds,
                'Multiple': comp_result.multiple,
                'Exit_Valuation': getattr(comp_result, 'exit_valuation', 0),
                'Exit_Ownership': getattr(comp_result, 'exit_ownership', 0),
                'Initial_Stage': comp_result.history[0]['stage'] if comp_result.history else '',
                'Final_Stage': comp_result.history[-1]['stage'] if comp_result.history else '',
                'Number_of_Rounds': len(comp_result.history),
                'Exit_or_progressed_to_next_stage': len(comp_result.history) > 1
            })
            
            # Extract round-by-round investment history
            for round_num, event in enumerate(comp_result.history, 1):
                all_investment_history.append({
                    'Simulation_Number': sim_num,
                    'Company_ID': comp_result.company_id,
                    'Round_Number': round_num,
                    'Time_Months': event['time'],
                    'Time_Years': event['time'] / 12,
                    'Event_Type': event['event'],
                    'Stage': event['stage'],
                    'PreMoney_Valuation': event.get('premoney_valuation', 0),
                    'Round_Investment': event.get('round_investment', 0),
                    'Round_Dilution': event.get('round_dilution', 0),
                    'PostMoney_Valuation': event.get('valuation', 0),
                    'Ownership_Pct': event.get('ownership', 0),
                    'Company_Outcome': comp_result.outcome,
                    'Final_Multiple': comp_result.multiple,
                    'Final_Exit_Proceeds': comp_result.exit_proceeds
                })
        
        # Extract cash flow transactions
        if not isinstance(gross_flows, list):
            print(f"Warning: Skipping cash flow processing for simulation {sim_num} - invalid format")
            continue
            
        for flow_num, (amount, time, company_id) in enumerate(gross_flows, 1):
            # Categorize cash flow types with descriptive labels
            if company_id == -2:
                flow_type = "Capital Call"
                flow_description = "Capital called from Limited Partners"
            elif company_id == -1:
                flow_type = "Management Fee"
                flow_description = "Annual management fee payment"
            elif company_id >= 0 and company_id < 9999:
                if amount < 0:
                    flow_type = "Investment"
                    flow_description = f"Investment in Company {company_id}"
                else:
                    flow_type = "Exit Proceeds"
                    flow_description = f"Exit proceeds from Company {company_id}"
            elif company_id == 9999:
                flow_type = "Cash Return"
                flow_description = "Unused cash returned to Limited Partners"
            else:
                flow_type = "Other"
                flow_description = "Other cash flow transaction"
            
            all_cash_flows_data.append({
                'Simulation_Number': sim_num,
                'Flow_Number': flow_num,
                'Time_Months': time,
                'Time_Years': time / 12,
                'Amount': amount,
                'Company_ID': company_id if company_id >= 0 else None,
                'Flow_Type': flow_type,
                'Flow_Description': flow_description,
                'Is_Inflow': amount > 0,
                'Is_Outflow': amount < 0,
                'Absolute_Amount': abs(amount)
            })
    
    # Create consolidated DataFrames
    df_companies = pd.DataFrame(all_company_data)
    df_history = pd.DataFrame(all_investment_history)
    df_flows = pd.DataFrame(all_cash_flows_data)
    
    # Sort for logical presentation
    df_companies = df_companies.sort_values(['Simulation_Number', 'Company_ID']).reset_index(drop=True)
    df_history = df_history.sort_values(['Simulation_Number', 'Company_ID', 'Round_Number']).reset_index(drop=True)
    df_flows = df_flows.sort_values(['Simulation_Number', 'Time_Months']).reset_index(drop=True)
    
    # Create formatted Excel workbook
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Main data sheets
        df_companies.to_excel(writer, sheet_name='Company_Summary', index=False)
        df_history.to_excel(writer, sheet_name='Investment_History', index=False)
        df_flows.to_excel(writer, sheet_name='Gross_Cash_Flows', index=False)
        
        # Optional detailed sheets
        if waterfall_log is not None and not waterfall_log.empty:
            waterfall_log_reordered = waterfall_log.copy().reset_index()
            waterfall_log_reordered['Year'] = waterfall_log_reordered.groupby('simulation_number').cumcount() + 1
            cols = ['simulation_number', 'Year'] + [col for col in waterfall_log_reordered.columns 
                                                  if col not in ['simulation_number', 'Year']]
            waterfall_log_reordered = waterfall_log_reordered[cols]
            waterfall_log_reordered.to_excel(writer, sheet_name='Waterfall_Details', index=False)
        
        if net_lp_flows_log is not None and not net_lp_flows_log.empty:
            net_lp_flows_log_reordered = net_lp_flows_log.copy()
            if 'simulation_number' in net_lp_flows_log_reordered.columns:
                cols = ['simulation_number'] + [col for col in net_lp_flows_log_reordered.columns 
                                               if col != 'simulation_number']
                net_lp_flows_log_reordered = net_lp_flows_log_reordered[cols]
            net_lp_flows_log_reordered.to_excel(writer, sheet_name='Net_LP_Flows', index=False)
        
        # Apply Excel formatting for readability
        workbook = writer.book
        
        # Format Company Summary sheet
        ws_companies = writer.sheets['Company_Summary']
        currency_cols_companies = ['Total_Invested', 'Exit_Proceeds', 'Exit_Valuation']
        for col_name in currency_cols_companies:
            if col_name in df_companies.columns:
                col_idx = df_companies.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df_companies) + 2):
                    ws_companies[f'{col_letter}{row}'].number_format = '$#,##0'
        
        # Format percentage columns
        pct_cols = ['Exit_Ownership']
        for col_name in pct_cols:
            if col_name in df_companies.columns:
                col_idx = df_companies.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df_companies) + 2):
                    ws_companies[f'{col_letter}{row}'].number_format = '0.0%'
        
        # Format multiple columns
        if 'Multiple' in df_companies.columns:
            mult_col_idx = df_companies.columns.get_loc('Multiple') + 1
            mult_col_letter = get_column_letter(mult_col_idx)
            for row in range(2, len(df_companies) + 2):
                ws_companies[f'{mult_col_letter}{row}'].number_format = '0.00"x"'
        
        # Format Investment History sheet
        ws_history = writer.sheets['Investment_History']
        currency_cols_history = ['PreMoney_Valuation', 'PostMoney_Valuation', 'Final_Exit_Proceeds']
        for col_name in currency_cols_history:
            if col_name in df_history.columns:
                col_idx = df_history.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df_history) + 2):
                    ws_history[f'{col_letter}{row}'].number_format = '$#,##0'
        
        if 'Ownership_Pct' in df_history.columns:
            own_col_idx = df_history.columns.get_loc('Ownership_Pct') + 1
            own_col_letter = get_column_letter(own_col_idx)
            for row in range(2, len(df_history) + 2):
                ws_history[f'{own_col_letter}{row}'].number_format = '0.0%'
        
        if 'Final_Multiple' in df_history.columns:
            mult_col_idx = df_history.columns.get_loc('Final_Multiple') + 1
            mult_col_letter = get_column_letter(mult_col_idx)
            for row in range(2, len(df_history) + 2):
                ws_history[f'{mult_col_letter}{row}'].number_format = '0.00"x"'
        
        # Format Gross Cash Flows sheet
        ws_flows = writer.sheets['Gross_Cash_Flows']
        currency_cols_flows = ['Amount', 'Absolute_Amount']
        for col_name in currency_cols_flows:
            if col_name in df_flows.columns:
                col_idx = df_flows.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df_flows) + 2):
                    ws_flows[f'{col_letter}{row}'].number_format = '$#,##0'
        
        # Format additional sheets if present
        if 'Waterfall_Details' in writer.sheets:
            ws_waterfall = writer.sheets['Waterfall_Details']
            currency_keywords = ['payment', 'contribution', 'distribution', 'proceeds', 'capital', 'cash']
            currency_cols_waterfall = [col for col in waterfall_log_reordered.columns 
                                     if any(keyword in col.lower() for keyword in currency_keywords)]
            for col_name in currency_cols_waterfall:
                if col_name in waterfall_log_reordered.columns:
                    col_idx = waterfall_log_reordered.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    for row in range(2, len(waterfall_log_reordered) + 2):
                        ws_waterfall[f'{col_letter}{row}'].number_format = '$#,##0'
        
        if 'Net_LP_Flows' in writer.sheets:
            ws_net_lp = writer.sheets['Net_LP_Flows']
            if 'amount' in net_lp_flows_log_reordered.columns:
                col_idx = net_lp_flows_log_reordered.columns.get_loc('amount') + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(net_lp_flows_log_reordered) + 2):
                    ws_net_lp[f'{col_letter}{row}'].number_format = '$#,##0'
    
    print(f"Excel export complete: {filename}")
    print(f"Summary:")
    print(f"  • Simulations: {len(all_results)}")
    print(f"  • Companies: {len(df_companies)} total across all simulations")
    print(f"  • Investment events: {len(df_history)}")
    print(f"  • Cash flows: {len(df_flows)}")
    if waterfall_log is not None:
        print(f"  • Waterfall records: {len(waterfall_log)}")
    if net_lp_flows_log is not None:
        print(f"  • Net LP flow records: {len(net_lp_flows_log)}")
    
    return df_companies, df_history, df_flows, waterfall_log, net_lp_flows_log